from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.exports.sip_workbook_contract import expected_result_formula


FIRST_ROW = 6
LAST_ROW = 517
LAST_BODY_ROW = 522

_THIN_BLACK = Side(style="thin", color="000000")
_BORDER = Border(
    left=_THIN_BLACK,
    right=_THIN_BLACK,
    top=_THIN_BLACK,
    bottom=_THIN_BLACK,
)
_BLUE = "4472C4"
_RED = "D9272E"
_ORANGE = "F06B2B"
_BODY_FILL = "FFF2DF"
_FIXED_CORE_MODIFIED = b"2026-07-31T00:00:00Z"
_TITLE_FONT = Font(name="Noto Serif CJK SC", size=20, bold=True, color="111111")
_BODY_FONT = Font(name="Noto Sans CJK SC", size=10, color="444444")
_BODY_BOLD_FONT = Font(name="Noto Sans CJK SC", size=10, bold=True, color="444444")
_ORANGE_BODY_FONT = Font(name="Noto Sans CJK SC", size=10, color=_ORANGE)
_RESULT_FONT = Font(name="Noto Sans CJK SC", size=10, bold=True, color="444444")
_NOTE_FONT = Font(name="Noto Sans CJK SC", size=10.5, bold=True, color=_RED)
_HEADER_FONT = Font(name="Noto Sans CJK SC", size=10.5, bold=True, color="FFFFFF")
_TYPE_COLORS = {
    "线性": "E5334E",
    "直径": "178BFF",
    "半径": "22B14C",
    "粗糙度": "C23ACF",
    "角度": "F39C3D",
    "螺纹": "009688",
    "技术要求": "6B7280",
    "复合": "B7791F",
}


def _normalize_archive(path: Path) -> None:
    members: dict[str, bytes] = {}
    with ZipFile(path, "r") as source:
        for name in source.namelist():
            members[name] = source.read(name)
    core_xml = members.get("docProps/core.xml")
    if core_xml is None:
        raise ValueError("workbook core properties are missing")
    members["docProps/core.xml"], replacements = re.subn(
        rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
        rb"\g<1>" + _FIXED_CORE_MODIFIED + rb"\g<2>",
        core_xml,
    )
    if replacements != 1:
        raise ValueError("workbook core modified timestamp is missing")
    normalized = path.with_suffix(".normalized.xlsx")
    with ZipFile(normalized, "w", compression=ZIP_DEFLATED) as target:
        for name in sorted(members):
            info = ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            target.writestr(info, members[name])
    normalized.replace(path)


def _style_range(sheet, cell_range: str, *, fill=None, font=None) -> None:
    for row in sheet[cell_range]:
        for cell in row:
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font


def _build_dimension_sheet(workbook) -> None:
    old_sheet = (
        workbook["SIP检验记录"]
        if "SIP检验记录" in workbook.sheetnames
        else workbook["尺寸质量检测表"]
    )
    workbook.remove(old_sheet)
    sheet = workbook.create_sheet("尺寸质量检测表", 0)
    if "气泡图" not in workbook.sheetnames:
        raise ValueError("registered balloon image worksheet is missing")
    if workbook["气泡图"]["A1"].value != "气泡图 / BALLOONED DRAWING":
        raise ValueError("registered balloon image worksheet title changed")

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A6"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_area = "A1:I522"
    sheet.print_title_rows = "1:5"
    sheet.sheet_properties.outlinePr.summaryBelow = True
    sheet.protection.sheet = True

    for column, width in {
        "A": 8,
        "B": 8,
        "C": 11,
        "D": 14,
        "E": 16,
        "F": 13,
        "G": 13,
        "H": 13,
        "I": 15,
    }.items():
        sheet.column_dimensions[column].width = width
    for row, height in {1: 36, 2: 24, 3: 24, 4: 28, 5: 27, 518: 24}.items():
        sheet.row_dimensions[row].height = height

    sheet.merge_cells("A1:I1")
    sheet["A1"] = "机械图纸尺寸质量检测表"
    sheet["A1"].font = _TITLE_FONT
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for cell_range in ("B2:D2", "F2:G2", "B3:D3", "F3:G3", "A4:I4"):
        sheet.merge_cells(cell_range)
    labels = {
        "A2": "文件名",
        "E2": "检测日期",
        "H2": "带公差",
        "A3": "总页数",
        "E3": "检验项总数",
        "H3": "单位",
    }
    for address, value in labels.items():
        sheet[address] = value
        sheet[address].font = _BODY_FONT
    sheet["A4"] = "【未注公差标准】未确认"
    sheet["A4"].font = _NOTE_FONT
    sheet["A4"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "编号",
        "页码",
        "类型",
        "基本尺寸",
        "公差",
        "上限",
        "下限",
        "检测值",
        "结果判定",
    ]
    header_fill = PatternFill("solid", fgColor=_BLUE)
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(5, index, header)
        cell.fill = header_fill
        cell.font = _HEADER_FONT

    _style_range(sheet, "A1:I5")
    for row in range(FIRST_ROW, LAST_ROW + 1):
        sheet.row_dimensions[row].height = 24
        for column in "ABCDEFGHI":
            cell = sheet[f"{column}{row}"]
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if column in "EFGHI":
                cell.fill = PatternFill("solid", fgColor=_BODY_FILL)
            if column in "ABCD":
                cell.font = _BODY_FONT
            elif column in "EFGH":
                cell.font = _ORANGE_BODY_FONT
            else:
                cell.font = _RESULT_FONT
        for column in "ABCDEFGH":
            sheet[f"{column}{row}"].protection = Protection(locked=False)
        sheet[f"H{row}"].value = None
        sheet[f"I{row}"] = expected_result_formula(row)
        sheet[f"I{row}"].protection = Protection(locked=True)

    for label, color in _TYPE_COLORS.items():
        sheet.conditional_formatting.add(
            f"C{FIRST_ROW}:C{LAST_ROW}",
            FormulaRule(
                formula=[f'$C{FIRST_ROW}="{label}"'],
                fill=PatternFill("solid", fgColor=color),
                font=Font(name="Noto Sans CJK SC", size=10, bold=True, color="FFFFFF"),
            ),
        )
    sheet.conditional_formatting.add(
        f"I{FIRST_ROW}:I{LAST_ROW}",
        FormulaRule(
            formula=[f'$I{FIRST_ROW}="OK"'],
            fill=PatternFill("solid", fgColor="C6EFCE"),
        ),
    )
    sheet.conditional_formatting.add(
        f"I{FIRST_ROW}:I{LAST_ROW}",
        FormulaRule(
            formula=[f'$I{FIRST_ROW}="NG"'],
            fill=PatternFill("solid", fgColor="FFC7CE"),
        ),
    )

    sheet.merge_cells("A518:I518")
    sheet["A518"] = "备注：检测值由质检人员填写，结果判定由受控公式自动计算。"
    sheet["A518"].font = _BODY_BOLD_FONT
    sheet["A518"].alignment = Alignment(horizontal="left", vertical="center")
    for cell_range, label in (
        ("A519:B522", "检验员："),
        ("C519:D522", "复核："),
        ("E519:F522", "批准："),
        ("G519:I522", "日期："),
    ):
        sheet.merge_cells(cell_range)
        cell = sheet[cell_range.split(":")[0]]
        cell.value = label
        cell.font = _BODY_BOLD_FONT
        cell.alignment = Alignment(horizontal="left", vertical="top")
    _style_range(sheet, "A518:I522")
    for row in range(518, LAST_BODY_ROW + 1):
        for column in "ABCDEFGHI":
            sheet[f"{column}{row}"].protection = Protection(locked=True)


def build_template(source_path: Path, target: Path) -> None:
    workbook = load_workbook(source_path, data_only=False)
    try:
        _build_dimension_sheet(workbook)
        workbook.properties.created = datetime(2026, 7, 31, tzinfo=timezone.utc)
        workbook.properties.modified = datetime(2026, 7, 31, tzinfo=timezone.utc)
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        target.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(target)
    finally:
        workbook.close()
    _normalize_archive(target)


def write_mapping(template_path: Path, mapping_path: Path) -> None:
    mapping = {
        "template_id": "sip-v1",
        "template_version": "3",
        "template_sha256": hashlib.sha256(template_path.read_bytes()).hexdigest(),
        "mapping_version": "3",
        "sheet": "尺寸质量检测表",
        "capacity": {"first_row": FIRST_ROW, "last_row": LAST_ROW},
        "metadata_cells": {
            "source_filename": "B2",
            "inspection_date": "F2",
            "toleranced_count": "I2",
            "page_count": "B3",
            "detail_count": "F3",
            "unit": "I3",
            "general_tolerance_note": "A4",
        },
        "detail_columns": {
            "number": "A",
            "source_page": "B",
            "type_label": "C",
            "basic_size": "D",
            "tolerance": "E",
            "upper_limit": "F",
            "lower_limit": "G",
        },
        "measurement_column": "H",
        "result_column": "I",
        "image_sheet": "气泡图",
        "image_anchor": "B2",
        "protected_ranges": [
            "A1:I1",
            "A2",
            "E2",
            "H2",
            "A3",
            "E3",
            "H3",
            "A5:I5",
            "I6:I517",
            "A518:I518",
        ],
        "signoff_ranges": ["A519:B522", "C519:D522", "E519:F522", "G519:I522"],
    }
    mapping_path.parent.mkdir(parents=True, exist_ok=True)
    mapping_path.write_text(
        json.dumps(mapping, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--target", type=Path, required=True)
    parser.add_argument("--mapping", type=Path, required=True)
    args = parser.parse_args()
    build_template(args.source, args.target)
    write_mapping(args.target, args.mapping)


if __name__ == "__main__":
    main()
