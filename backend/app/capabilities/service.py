from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.exports.template_registry import (
    AssetHashMismatch,
    InvalidTemplateRegistration,
    MappingHashMismatch,
    TemplateRegistration,
    file_sha256,
    load_template_registration,
)


APPROVED_BALLOON_FONT_SHA256 = (
    "ae7b7855e115a5966d8b1b3f80f254ccc117ec86f9965e202ee2940453837280"
)
APPROVED_BALLOON_FONT_LICENSE_SHA256 = (
    "63d3ba759d12804c5b31a9d5940d855c1820d1f5999e6b0872eb1c7ff045fbc9"
)

class CapabilityUnavailable(RuntimeError):
    def __init__(self, code: str, detail: str) -> None:
        super().__init__(detail)
        self.code = code
        self.detail = detail


def _configured(value: bool | str | None) -> bool:
    if isinstance(value, str):
        return bool(value.strip())
    return value is True


class ProcessingPreflight:
    def __init__(
        self,
        storage: Any,
        redis_client: Any,
        celery_application: Any | None = None,
        *,
        ocr_configured: bool | str | None,
        vision_configured: bool | str | None,
    ) -> None:
        if celery_application is None:
            from app.celery_app import celery_app

            celery_application = celery_app
        self._storage = storage
        self._redis = redis_client
        self._celery = celery_application
        self._ocr_configured = ocr_configured
        self._vision_configured = vision_configured

    def check(self) -> None:
        try:
            self._storage.probe()
        except Exception as exc:
            raise CapabilityUnavailable(
                "storage_unavailable",
                "shared storage probe failed",
            ) from exc

        try:
            if self._redis.ping() is not True:
                raise RuntimeError("Redis ping did not return true")
        except Exception as exc:
            raise CapabilityUnavailable(
                "redis_unavailable",
                "Redis capability check failed",
            ) from exc

        try:
            workers = self._celery.control.inspect(timeout=1).ping()
            if not isinstance(workers, dict) or not workers:
                raise RuntimeError("no Celery worker responded")
        except Exception as exc:
            raise CapabilityUnavailable(
                "celery_worker_unavailable",
                "Celery worker capability check failed",
            ) from exc

        if not _configured(self._ocr_configured):
            raise CapabilityUnavailable(
                "ocr_provider_unavailable",
                "OCR Provider configuration is unavailable",
            )
        if not _configured(self._vision_configured):
            raise CapabilityUnavailable(
                "vision_provider_unavailable",
                "Vision Provider configuration is unavailable",
            )


class ExportPreflight:
    def __init__(
        self,
        *,
        template_path: Path,
        mapping_path: Path,
        font_path: Path,
        font_license_path: Path,
    ) -> None:
        self._template_path = template_path
        self._mapping_path = mapping_path
        self._font_path = font_path
        self._font_license_path = font_license_path

    def check(self) -> TemplateRegistration:
        self._require_file(
            self._template_path,
            "export_template_unavailable",
            "registered export template is unavailable",
        )
        self._require_file(
            self._mapping_path,
            "export_template_mapping_unavailable",
            "registered export template mapping is unavailable",
        )
        self._require_file(
            self._font_path,
            "export_font_unavailable",
            "registered balloon font is unavailable",
        )
        self._require_file(
            self._font_license_path,
            "export_font_license_unavailable",
            "registered balloon font license is unavailable",
        )

        try:
            registration = load_template_registration(
                self._template_path,
                self._mapping_path,
            )
        except MappingHashMismatch as exc:
            raise CapabilityUnavailable(
                "export_template_mapping_hash_mismatch",
                "registered export template mapping hash does not match",
            ) from exc
        except AssetHashMismatch as exc:
            raise CapabilityUnavailable(
                "export_template_hash_mismatch",
                "registered export template hash does not match",
            ) from exc
        except InvalidTemplateRegistration as exc:
            raise CapabilityUnavailable(
                "export_template_registration_invalid",
                "registered export template mapping is invalid",
            ) from exc

        try:
            workbook = load_workbook(
                self._template_path,
                read_only=True,
                data_only=False,
            )
        except (OSError, ValueError) as exc:
            raise CapabilityUnavailable(
                "export_template_invalid",
                "registered export template cannot be opened",
            ) from exc
        try:
            missing_sheets = {
                registration.sheet,
                registration.image_sheet,
            } - set(workbook.sheetnames)
        finally:
            workbook.close()
        if missing_sheets:
            raise CapabilityUnavailable(
                "export_template_sheet_missing",
                "registered export template sheet is missing",
            )

        try:
            font_sha256 = file_sha256(self._font_path)
        except OSError as exc:
            raise CapabilityUnavailable(
                "export_font_unavailable",
                "registered balloon font is unavailable",
            ) from exc
        if font_sha256 != APPROVED_BALLOON_FONT_SHA256:
            raise CapabilityUnavailable(
                "export_font_hash_mismatch",
                "registered balloon font hash does not match",
            )

        try:
            font_license_sha256 = file_sha256(self._font_license_path)
        except OSError as exc:
            raise CapabilityUnavailable(
                "export_font_license_unavailable",
                "registered balloon font license is unavailable",
            ) from exc
        if font_license_sha256 != APPROVED_BALLOON_FONT_LICENSE_SHA256:
            raise CapabilityUnavailable(
                "export_font_license_hash_mismatch",
                "registered balloon font license hash does not match",
            )
        return registration

    @staticmethod
    def _require_file(path: Path, code: str, detail: str) -> None:
        if not path.is_file():
            raise CapabilityUnavailable(code, detail)
