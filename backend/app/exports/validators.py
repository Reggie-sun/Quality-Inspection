from __future__ import annotations

from io import BytesIO
from typing import TypeAlias

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from app.exports.sip_workbook_contract import (
    NUMERIC_DETAIL_FIELDS,
    expected_result_formula,
)
from app.exports.template_registry import TemplateRegistration


RangeSnapshot: TypeAlias = tuple[
    tuple[str, object, int, tuple[str, ...]],
    ...,
]


def snapshot_registered_ranges(
    workbook,
    registration: TemplateRegistration,
) -> RangeSnapshot:
    sheet = workbook[registration.sheet]
    merged_ranges = tuple(sheet.merged_cells.ranges)
    cells: dict[str, tuple[str, object, int, tuple[str, ...]]] = {}
    for cell_range in registration.protected_ranges + registration.signoff_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                cell = sheet.cell(row=row, column=column)
                cells[cell.coordinate] = (
                    cell.coordinate,
                    cell.value,
                    cell.style_id,
                    tuple(
                        sorted(
                            str(merged)
                            for merged in merged_ranges
                            if cell.coordinate in merged
                        )
                    ),
                )
    return tuple(cells[coordinate] for coordinate in sorted(cells))


def validate_sip_workbook(
    content: bytes,
    registration: TemplateRegistration,
    *,
    protected_snapshot: RangeSnapshot,
    detail_count: int,
    source_page_count: int,
) -> None:
    try:
        workbook = load_workbook(BytesIO(content), data_only=False)
    except (OSError, ValueError) as exc:
        raise ValueError("generated workbook cannot be reopened") from exc
    try:
        missing_sheets = {
            registration.sheet,
            registration.image_sheet,
        } - set(workbook.sheetnames)
        if missing_sheets:
            raise ValueError("generated workbook sheet is missing")
        if snapshot_registered_ranges(workbook, registration) != protected_snapshot:
            raise ValueError("registered fixed or sign-off range changed")

        sheet = workbook[registration.sheet]
        for offset in range(detail_count):
            row = registration.first_row + offset
            for field, column in registration.detail_columns.items():
                cell = sheet[f"{column}{row}"]
                if field in NUMERIC_DETAIL_FIELDS:
                    if cell.value not in (None, "") and cell.data_type != "n":
                        raise ValueError("generated numeric detail cell is not numeric")
                elif cell.value is not None and cell.data_type != "s":
                    raise ValueError("generated text detail cell is not plain text")
                if cell.protection.locked:
                    raise ValueError("generated detail cell is not editable")

        for row in range(registration.first_row, registration.last_row + 1):
            measurement = sheet[f"{registration.measurement_column}{row}"]
            if measurement.value not in (None, "") or measurement.protection.locked:
                raise ValueError("measurement cell is not blank and editable")
            result = sheet[f"{registration.result_column}{row}"]
            if result.value != expected_result_formula(row):
                raise ValueError("trusted result formula changed")

        image_count = len(workbook[registration.image_sheet]._images)
        if image_count != source_page_count:
            raise ValueError("embedded image count does not match source pages")

        resaved = BytesIO()
        workbook.save(resaved)
    finally:
        workbook.close()

    try:
        reopened = load_workbook(BytesIO(resaved.getvalue()), data_only=False)
    except (OSError, ValueError) as exc:
        raise ValueError("generated workbook cannot be resaved and reopened") from exc
    try:
        if len(reopened[registration.image_sheet]._images) != source_page_count:
            raise ValueError("embedded images changed after workbook resave")
    finally:
        reopened.close()
