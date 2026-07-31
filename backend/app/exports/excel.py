from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.drawing.image import Image

from app.exports.template_registry import TemplateRegistration
from app.exports.sip_workbook_contract import (
    NUMERIC_DETAIL_FIELDS,
    NUMERIC_METADATA_FIELDS,
    TEXT_DETAIL_FIELDS,
    TEXT_METADATA_FIELDS,
    expected_result_formula,
)
from app.exports.validators import (
    snapshot_registered_ranges,
    validate_sip_workbook,
)


REQUIRED_METADATA_FIELDS = NUMERIC_METADATA_FIELDS | TEXT_METADATA_FIELDS
REQUIRED_DETAIL_FIELDS = NUMERIC_DETAIL_FIELDS | TEXT_DETAIL_FIELDS


class CapacityExceeded(RuntimeError):
    pass


def set_untrusted_text(cell: Cell, value: object) -> None:
    cell.value = "" if value is None else str(value)
    cell.data_type = "s"


def set_registered_value(cell: Cell, field: str, value: object) -> None:
    if value in (None, ""):
        cell.value = None
        return
    if field not in NUMERIC_DETAIL_FIELDS:
        set_untrusted_text(cell, value)
        return
    if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
        raise ValueError(f"{field} must be numeric")
    try:
        numeric = Decimal(str(value))
    except (InvalidOperation, ValueError) as error:
        raise ValueError(f"{field} must be numeric") from error
    if not numeric.is_finite():
        raise ValueError(f"{field} must be numeric")
    cell.value = int(numeric) if numeric == numeric.to_integral_value() else float(numeric)
    cell.data_type = "n"


def set_metadata_value(cell: Cell, field: str, value: object) -> None:
    if field not in NUMERIC_METADATA_FIELDS:
        set_untrusted_text(cell, value)
        return
    if isinstance(value, bool) or not isinstance(value, int) or value < 0:
        raise ValueError(f"{field} must be a non-negative integer")
    cell.value = value
    cell.data_type = "n"


def assert_capacity(
    registration: TemplateRegistration,
    detail_count: int,
) -> None:
    if detail_count > registration.capacity:
        raise CapacityExceeded(
            f"{detail_count} details exceed capacity {registration.capacity}"
        )


def render_sip_workbook(
    template_path: Path,
    registration: TemplateRegistration,
    metadata: dict[str, object],
    reviewed_items: list[dict[str, object]],
    page_images: list[Path],
) -> bytes:
    if set(registration.metadata_cells) != REQUIRED_METADATA_FIELDS:
        raise ValueError("fixed metadata mapping is incomplete")
    if set(registration.detail_columns) != REQUIRED_DETAIL_FIELDS:
        raise ValueError("fixed detail mapping is incomplete")
    assert_capacity(registration, len(reviewed_items))

    book = load_workbook(template_path, data_only=False)
    try:
        missing_sheets = {
            registration.sheet,
            registration.image_sheet,
        } - set(book.sheetnames)
        if missing_sheets:
            raise ValueError("registered workbook sheet is missing")
        protected_snapshot = snapshot_registered_ranges(book, registration)
        sheet = book[registration.sheet]
        for field, address in registration.metadata_cells.items():
            set_metadata_value(sheet[address], field, metadata[field])

        for offset, item in enumerate(reviewed_items):
            row = registration.first_row + offset
            for field, column in registration.detail_columns.items():
                set_registered_value(sheet[f"{column}{row}"], field, item.get(field, ""))
            measurement = sheet[f"{registration.measurement_column}{row}"]
            if measurement.value not in (None, "") or measurement.protection.locked:
                raise ValueError("measurement cell is not blank and editable")
            result = sheet[f"{registration.result_column}{row}"]
            if result.value != expected_result_formula(row):
                raise ValueError("trusted result formula changed")

        image_sheet = book[registration.image_sheet]
        anchor_cell = image_sheet[registration.image_anchor]
        for page_index, image_path in enumerate(page_images):
            image = Image(image_path)
            image.anchor = (
                f"{anchor_cell.column_letter}{anchor_cell.row + page_index * 45}"
            )
            image_sheet.add_image(image)

        output = BytesIO()
        book.save(output)
        content = output.getvalue()
    finally:
        book.close()

    validate_sip_workbook(
        content,
        registration,
        protected_snapshot=protected_snapshot,
        detail_count=len(reviewed_items),
        source_page_count=len(page_images),
    )
    return content
