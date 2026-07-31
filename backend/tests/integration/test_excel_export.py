from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from PIL import Image as PillowImage

from app.exports.excel import render_sip_workbook
from app.exports.template_registry import TemplateRegistration, load_template_registration
from app.exports.validators import snapshot_registered_ranges, validate_sip_workbook


def _approved_assets() -> tuple[Path, Path, TemplateRegistration]:
    backend_root = Path(__file__).resolve().parents[2]
    template_path = backend_root / "assets/templates/sip-v1.xlsx"
    mapping_path = backend_root / "assets/templates/sip-v1.mapping.json"
    return (
        template_path,
        mapping_path,
        load_template_registration(template_path, mapping_path),
    )


def _metadata() -> dict[str, object]:
    return {
        "source_filename": "drawing.pdf",
        "inspection_date": "2026-07-31 10:30",
        "toleranced_count": 1,
        "page_count": 2,
        "detail_count": 2,
        "unit": "mm / 按项目",
        "general_tolerance_note": "【未注公差标准】未确认",
    }


def _reviewed_items() -> list[dict[str, object]]:
    return [
        {
            "number": 1,
            "source_page": 1,
            "type_label": "线性",
            "basic_size": "500",
            "tolerance": "±0.2",
            "upper_limit": Decimal("500.2"),
            "lower_limit": Decimal("499.8"),
        },
        {
            "number": "",
            "source_page": 2,
            "type_label": "技术要求",
            "basic_size": "去毛刺",
            "tolerance": "",
            "upper_limit": "",
            "lower_limit": "",
        },
    ]


def _page_images(tmp_path: Path) -> tuple[list[Path], list[tuple[int, int, int]]]:
    colors = [(240, 10, 10), (10, 20, 230)]
    paths: list[Path] = []
    for page_number, color in enumerate(colors, start=1):
        path = tmp_path / f"page-{page_number}.png"
        PillowImage.new("RGB", (24, 16), color=color).save(path)
        paths.append(path)
    return paths, colors


def _render(tmp_path: Path) -> tuple[bytes, TemplateRegistration, Path]:
    template_path, _, registration = _approved_assets()
    page_images, _ = _page_images(tmp_path)
    return (
        render_sip_workbook(
            template_path,
            registration,
            _metadata(),
            _reviewed_items(),
            page_images,
        ),
        registration,
        template_path,
    )


def _registered_range_snapshot(
    workbook,
    registration: TemplateRegistration,
) -> dict[str, tuple[object, int, tuple[str, ...]]]:
    sheet = workbook[registration.sheet]
    merged_ranges = tuple(sheet.merged_cells.ranges)
    snapshot: dict[str, tuple[object, int, tuple[str, ...]]] = {}
    for cell_range in registration.protected_ranges + registration.signoff_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                cell = sheet.cell(row=row, column=column)
                membership = tuple(
                    sorted(
                        str(merged)
                        for merged in merged_ranges
                        if cell.coordinate in merged
                    )
                )
                snapshot[cell.coordinate] = (cell.value, cell.style_id, membership)
    return snapshot


def test_v3_renderer_writes_numeric_cells_and_preserves_template_contract(
    tmp_path: Path,
) -> None:
    """Catches renderer regressions that stringify v3 numbers or overwrite H/I."""
    content, registration, _ = _render(tmp_path)

    workbook = load_workbook(BytesIO(content), data_only=False)
    try:
        sheet = workbook[registration.sheet]
        assert sheet["A6"].value == 1
        assert sheet["B6"].value == 1
        assert sheet["C6"].value == "线性"
        assert sheet["D6"].value == "500"
        assert sheet["E6"].value == "±0.2"
        assert sheet["F6"].value == 500.2
        assert sheet["G6"].value == 499.8
        assert sheet["A6"].data_type == "n"
        assert sheet["B6"].data_type == "n"
        assert sheet["F6"].data_type == "n"
        assert sheet["G6"].data_type == "n"
        assert sheet["I2"].data_type == "n"
        assert sheet["B3"].data_type == "n"
        assert sheet["F3"].data_type == "n"
        assert sheet["A7"].value is None
        assert sheet["H6"].value is None
        assert sheet["H6"].protection.locked is False
        assert sheet["H517"].value is None
        assert sheet["H517"].protection.locked is False
        assert sheet["I6"].value == (
            '=IF(H6="","",IF(OR(F6="",G6=""),"",'
            'IF(AND(ISNUMBER(H6),H6<=F6,H6>=G6),"OK","NG")))'
        )
        assert sheet["I517"].value == (
            '=IF(H517="","",IF(OR(F517="",G517=""),"",'
            'IF(AND(ISNUMBER(H517),H517<=F517,H517>=G517),"OK","NG")))'
        )
        assert sheet["A1"].value == "机械图纸尺寸质量检测表"
        assert sheet["A4"].font.color.rgb[-6:] == "D9272E"
        assert sheet["A5"].fill.fgColor.rgb[-6:] == "4472C4"
        assert str(sheet.print_area) == "'尺寸质量检测表'!$A$1:$I$522"
        assert sheet.print_title_rows == "$1:$5"
        assert sheet.page_setup.orientation == "landscape"
        conditional_fills: dict[str, str] = {}
        for conditional in sheet.conditional_formatting:
            for rule in sheet.conditional_formatting[conditional]:
                if rule.formula and rule.dxf is not None and rule.dxf.fill is not None:
                    conditional_fills[rule.formula[0]] = rule.dxf.fill.fgColor.rgb[-6:]
        assert conditional_fills['$C6="线性"'] == "E5334E"
        assert conditional_fills['$C6="直径"'] == "178BFF"
        assert conditional_fills['$C6="半径"'] == "22B14C"
        assert conditional_fills['$C6="粗糙度"'] == "C23ACF"
    finally:
        workbook.close()


def test_renderer_keeps_formula_like_text_as_plain_text(tmp_path: Path) -> None:
    """Catches a text cell being promoted to an Excel formula by untrusted input."""
    template_path, _, registration = _approved_assets()
    page_images, _ = _page_images(tmp_path)
    rows = _reviewed_items()
    rows[0]["basic_size"] = '=HYPERLINK("https://invalid")'

    content = render_sip_workbook(
        template_path,
        registration,
        _metadata(),
        rows,
        page_images,
    )

    workbook = load_workbook(BytesIO(content), data_only=False)
    try:
        assert workbook[registration.sheet]["D6"].data_type == "s"
    finally:
        workbook.close()


@pytest.mark.parametrize("invalid_value", ["500.2", True, float("nan")])
def test_renderer_rejects_non_trusted_numeric_values(
    tmp_path: Path,
    invalid_value: object,
) -> None:
    """Catches numeric detail inputs that are strings, bools, or non-finite values."""
    template_path, _, registration = _approved_assets()
    page_images, _ = _page_images(tmp_path)
    rows = _reviewed_items()
    rows[0]["upper_limit"] = invalid_value

    with pytest.raises(ValueError, match="upper_limit must be numeric"):
        render_sip_workbook(
            template_path,
            registration,
            _metadata(),
            rows,
            page_images,
        )


def test_validator_rejects_a_mutated_trusted_result_formula(tmp_path: Path) -> None:
    """Catches a protected result cell whose formula was changed in the template."""
    template_path, _, registration = _approved_assets()
    workbook = load_workbook(template_path, data_only=False)
    try:
        workbook[registration.sheet]["I6"] = "=1+1"
        protected_snapshot = snapshot_registered_ranges(workbook, registration)
        content = BytesIO()
        workbook.save(content)
    finally:
        workbook.close()

    with pytest.raises(ValueError, match="trusted result formula changed"):
        validate_sip_workbook(
            content.getvalue(),
            registration,
            protected_snapshot=protected_snapshot,
            detail_count=0,
            source_page_count=0,
        )


def test_all_ballooned_pages_are_embedded_in_order(tmp_path: Path) -> None:
    """P0-EXP-005 embeds formal ballooned-PDF pages in source-page order."""
    content, registration, _ = _render(tmp_path)
    _, expected_colors = _page_images(tmp_path)
    workbook = load_workbook(BytesIO(content), data_only=False)
    try:
        images = workbook[registration.image_sheet]._images
        assert [
            (image.anchor._from.row, image.anchor._from.col) for image in images
        ] == [(1, 1), (46, 1)]
        actual_colors = []
        for image in images:
            with PillowImage.open(BytesIO(image._data())) as embedded:
                actual_colors.append(embedded.convert("RGB").getpixel((0, 0)))
        assert actual_colors == expected_colors
    finally:
        workbook.close()


def test_fixed_and_signoff_ranges_are_preserved(tmp_path: Path) -> None:
    """P0-EXP-007F preserves registered fixed/sign-off values and styles."""
    content, registration, template_path = _render(tmp_path)
    original = load_workbook(template_path, data_only=False)
    generated = load_workbook(BytesIO(content), data_only=False)
    try:
        assert _registered_range_snapshot(generated, registration) == (
            _registered_range_snapshot(original, registration)
        )
    finally:
        generated.close()
        original.close()
