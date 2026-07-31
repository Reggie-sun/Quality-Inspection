from __future__ import annotations

import json
import importlib.util
import types
from datetime import datetime, timezone
from hashlib import sha256
from pathlib import Path
from shutil import copyfile
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.writer import excel as openpyxl_excel_writer

from app.exports.template_registry import (
    AssetHashMismatch,
    InvalidTemplateRegistration,
    load_template_registration,
)


def _complete_mapping(template_bytes: bytes) -> dict[str, object]:
    return {
        "template_id": "sip-v1",
        "template_version": "3",
        "template_sha256": sha256(template_bytes).hexdigest(),
        "mapping_version": "3",
        "sheet": "尺寸质量检测表",
        "capacity": {"first_row": 6, "last_row": 517},
        "metadata_cells": {
            "source_filename": "B2",
            "inspection_date": "F2",
            "toleranced_count": "I2",
            "page_count": "B3",
            "detail_count": "F3",
            "unit": "I3",
            "general_tolerance_note": "A4",
        },
        "detail_columns": {
            "number": "A",
            "source_page": "B",
            "type_label": "C",
            "basic_size": "D",
            "tolerance": "E",
            "upper_limit": "F",
            "lower_limit": "G",
        },
        "measurement_column": "H",
        "result_column": "I",
        "image_sheet": "气泡图",
        "image_anchor": "B2",
        "protected_ranges": [
            "A1:I1",
            "A2",
            "E2",
            "H2",
            "A3",
            "E3",
            "H3",
            "A5:I5",
            "I6:I517",
            "A518:I518",
        ],
        "signoff_ranges": [
            "A519:B522",
            "C519:D522",
            "E519:F522",
            "G519:I522",
        ],
    }


def _write_registration(
    tmp_path: Path,
    template_bytes: bytes,
    mapping: dict[str, object],
) -> tuple[Path, Path]:
    template_path = tmp_path / "sip-v1.xlsx"
    mapping_path = tmp_path / "sip-v1.mapping.json"
    template_path.write_bytes(template_bytes)
    mapping_path.write_text(json.dumps(mapping), encoding="utf-8")
    return template_path, mapping_path


def _copy_approved_registration(tmp_path: Path) -> tuple[Path, Path]:
    backend_root = Path(__file__).resolve().parents[3]
    template_path = tmp_path / "sip-v1.xlsx"
    mapping_path = tmp_path / "sip-v1.mapping.json"
    copyfile(backend_root / "assets/templates/sip-v1.xlsx", template_path)
    copyfile(backend_root / "assets/templates/sip-v1.mapping.json", mapping_path)
    return template_path, mapping_path


def _load_template_builder():
    backend_root = Path(__file__).resolve().parents[3]
    builder_path = backend_root / "scripts/build_sip_template_v3.py"
    spec = importlib.util.spec_from_file_location("sip_template_v3_builder", builder_path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def test_p0_exp_001_loads_the_approved_v3_template_registration() -> None:
    """Catches a registered workbook that retains the v2 layout or row columns."""
    backend_root = Path(__file__).resolve().parents[3]
    registration = load_template_registration(
        backend_root / "assets/templates/sip-v1.xlsx",
        backend_root / "assets/templates/sip-v1.mapping.json",
    )

    assert registration.template_id == "sip-v1"
    assert registration.template_version == "3"
    assert registration.mapping_version == "3"
    assert registration.sheet == "尺寸质量检测表"
    assert registration.capacity == 512
    assert registration.first_row == 6
    assert registration.last_row == 517
    assert registration.measurement_column == "H"
    assert registration.result_column == "I"
    assert registration.image_sheet == "气泡图"
    assert registration.image_anchor == "B2"


def test_p0_exp_003_approved_v3_template_has_fixed_print_capacity() -> None:
    """Catches a v3 workbook that moves its fixed body or corrupts H/I ownership."""
    backend_root = Path(__file__).resolve().parents[3]
    template_path = backend_root / "assets/templates/sip-v1.xlsx"
    registration = load_template_registration(
        template_path,
        backend_root / "assets/templates/sip-v1.mapping.json",
    )
    workbook = load_workbook(template_path, data_only=False)
    try:
        sheet = workbook[registration.sheet]
        assert (registration.first_row, registration.last_row) == (6, 517)
        assert sheet.max_row == 522
        assert str(sheet.print_area) == "'尺寸质量检测表'!$A$1:$I$522"
        assert sheet.print_title_rows == "$1:$5"
        assert registration.protected_ranges == (
            "A1:I1",
            "A2",
            "E2",
            "H2",
            "A3",
            "E3",
            "H3",
            "A5:I5",
            "I6:I517",
            "A518:I518",
        )
        assert registration.signoff_ranges == (
            "A519:B522",
            "C519:D522",
            "E519:F522",
            "G519:I522",
        )
        assert all(
            sheet[f"H{row}"].value is None
            and sheet[f"H{row}"].protection.locked is False
            and sheet[f"I{row}"].value
            == (
                f'=IF(H{row}="","",IF(OR(F{row}="",G{row}=""),"",'
                f'IF(AND(ISNUMBER(H{row}),H{row}<=F{row},H{row}>=G{row}),"OK","NG")))'
            )
            and sheet[f"I{row}"].protection.locked is True
            for row in range(registration.first_row, registration.last_row + 1)
        )
        assert all(
            sheet[f"{column}{row}"].protection.locked is True
            for row in range(518, 523)
            for column in "ABCDEFGHI"
        )
        assert sheet["A1"].font.name == "Noto Serif CJK SC"
        assert sheet["A1"].font.sz == 20
        assert sheet["A5"].font.name == "Noto Sans CJK SC"
        assert sheet["A5"].font.sz == 10.5
        assert sheet["A5"].font.color.rgb[-6:] == "FFFFFF"
        assert {column: sheet.column_dimensions[column].width for column in "ABCDEFGHI"} == {
            "A": 8,
            "B": 8,
            "C": 11,
            "D": 14,
            "E": 16,
            "F": 13,
            "G": 13,
            "H": 13,
            "I": 15,
        }
        assert {row: sheet.row_dimensions[row].height for row in range(1, 7)} == {
            1: 36,
            2: 24,
            3: 24,
            4: 28,
            5: 27,
            6: 24,
        }
        assert sheet["A6"].font.name == "Noto Sans CJK SC"
        assert sheet["A6"].font.sz == 10
        assert sheet["A6"].font.color.rgb[-6:] == "444444"
        assert sheet["E6"].fill.fgColor.rgb[-6:] == "FFF2DF"
        assert sheet["E6"].font.color.rgb[-6:] == "F06B2B"
        assert sheet["H6"].fill.fgColor.rgb[-6:] == "FFF2DF"
        assert sheet["H6"].font.color.rgb[-6:] == "F06B2B"
        assert sheet["I6"].fill.fgColor.rgb[-6:] == "FFF2DF"
        assert sheet["I6"].font.bold is True
        type_rules = {}
        for conditional in sheet.conditional_formatting:
            for rule in sheet.conditional_formatting[conditional]:
                if rule.formula and rule.dxf is not None and rule.dxf.font is not None:
                    type_rules[rule.formula[0]] = rule.dxf.font
        assert type_rules['$C6="线性"'].bold is True
        assert type_rules['$C6="线性"'].color.rgb[-6:] == "FFFFFF"
    finally:
        workbook.close()


def test_builder_normalizes_core_properties_across_distinct_save_times(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Catches a builder whose XLSX bytes drift with openpyxl's writer clock."""
    backend_root = Path(__file__).resolve().parents[3]
    source = backend_root / "assets/templates/sip-v1.xlsx"
    builder = _load_template_builder()

    class WriterClock:
        current = datetime(2026, 7, 31, 1, tzinfo=timezone.utc)

        @classmethod
        def now(cls, tz=None):
            return cls.current if tz is not None else cls.current.replace(tzinfo=None)

    monkeypatch.setattr(
        openpyxl_excel_writer,
        "datetime",
        types.SimpleNamespace(datetime=WriterClock, timezone=timezone),
    )
    first_template = tmp_path / "first.xlsx"
    first_mapping = tmp_path / "first.mapping.json"
    builder.build_template(source, first_template)
    builder.write_mapping(first_template, first_mapping)

    WriterClock.current = datetime(2026, 7, 31, 2, tzinfo=timezone.utc)
    second_template = tmp_path / "second.xlsx"
    second_mapping = tmp_path / "second.mapping.json"
    builder.build_template(source, second_template)
    builder.write_mapping(second_template, second_mapping)

    assert first_template.read_bytes() == second_template.read_bytes()
    assert first_mapping.read_bytes() == second_mapping.read_bytes()
    with ZipFile(first_template) as archive:
        core_xml = archive.read("docProps/core.xml")
    assert b"2026-07-31T00:00:00Z" in core_xml


def test_p0_exp_001_registry_rejects_template_hash_drift(tmp_path: Path) -> None:
    """P0-EXP-001 fails closed when registered bytes no longer match their hash."""
    template_path, mapping_path = _copy_approved_registration(tmp_path)
    template_path.write_bytes(b"changed-template-bytes")

    with pytest.raises(AssetHashMismatch, match="template hash drift"):
        load_template_registration(template_path, mapping_path)


def test_p0_exp_001_rejects_joint_template_and_mapping_replacement(
    tmp_path: Path,
) -> None:
    """P0-EXP-001 anchors identity outside the two mutually editable assets."""
    _, mapping_path = _copy_approved_registration(tmp_path)
    template_path = tmp_path / "replacement.xlsx"
    workbook = Workbook()
    workbook.active.title = "SIP检验记录"
    workbook.create_sheet("气泡图")
    workbook.save(template_path)
    mapping = json.loads(mapping_path.read_text(encoding="utf-8"))
    mapping["template_sha256"] = sha256(template_path.read_bytes()).hexdigest()
    mapping_path.write_text(json.dumps(mapping, ensure_ascii=False), encoding="utf-8")

    with pytest.raises(AssetHashMismatch, match="mapping hash drift"):
        load_template_registration(template_path, mapping_path)


def test_p0_exp_001_registry_rejects_missing_fixed_mapping_field(
    tmp_path: Path,
) -> None:
    """P0-EXP-001 refuses an incomplete fixed-field registration."""
    template_bytes = b"approved-template-bytes"
    mapping = _complete_mapping(template_bytes)
    detail_columns = dict(mapping["detail_columns"])
    detail_columns.pop("source_page")
    mapping["detail_columns"] = detail_columns
    template_path, mapping_path = _write_registration(
        tmp_path,
        template_bytes,
        mapping,
    )

    with pytest.raises(
        InvalidTemplateRegistration,
        match="complete fixed-field mapping",
    ):
        load_template_registration(template_path, mapping_path)


def test_registry_rejects_a_registration_without_a_measurement_column(
    tmp_path: Path,
) -> None:
    """Catches a v3 mapping that cannot identify the editable measurement column."""
    template_bytes = b"approved-template-bytes"
    mapping = _complete_mapping(template_bytes)
    mapping.pop("measurement_column")
    template_path, mapping_path = _write_registration(
        tmp_path,
        template_bytes,
        mapping,
    )

    with pytest.raises(
        InvalidTemplateRegistration,
        match="measurement_column must be a non-empty string",
    ):
        load_template_registration(template_path, mapping_path)


def test_registry_rejects_result_column_that_overlaps_measurement_column(
    tmp_path: Path,
) -> None:
    """Catches a v3 mapping that would overwrite hand-entered measurements."""
    template_bytes = b"approved-template-bytes"
    mapping = _complete_mapping(template_bytes)
    mapping["result_column"] = "H"
    template_path, mapping_path = _write_registration(
        tmp_path,
        template_bytes,
        mapping,
    )

    with pytest.raises(
        InvalidTemplateRegistration,
        match="measurement_column and result_column must not overlap",
    ):
        load_template_registration(template_path, mapping_path)


def test_registry_rejects_detail_column_that_overlaps_result_formula(
    tmp_path: Path,
) -> None:
    """Catches a v3 mapping that would replace the trusted result formula."""
    template_bytes = b"approved-template-bytes"
    mapping = _complete_mapping(template_bytes)
    detail_columns = dict(mapping["detail_columns"])
    detail_columns["upper_limit"] = "I"
    mapping["detail_columns"] = detail_columns
    template_path, mapping_path = _write_registration(
        tmp_path,
        template_bytes,
        mapping,
    )

    with pytest.raises(
        InvalidTemplateRegistration,
        match="detail_columns must not overlap measurement or result columns",
    ):
        load_template_registration(template_path, mapping_path)


def test_p0_exp_001_registry_rejects_any_template_id_other_than_sip_v1(
    tmp_path: Path,
) -> None:
    """P0-EXP-001 keeps P0 on its one approved template identity."""
    template_bytes = b"approved-template-bytes"
    mapping = _complete_mapping(template_bytes)
    mapping["template_id"] = "unregistered-template"
    template_path, mapping_path = _write_registration(
        tmp_path,
        template_bytes,
        mapping,
    )

    with pytest.raises(InvalidTemplateRegistration, match="sip-v1"):
        load_template_registration(template_path, mapping_path)
