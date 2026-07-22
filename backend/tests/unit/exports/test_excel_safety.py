from __future__ import annotations

from openpyxl import Workbook, load_workbook

from app.exports.excel import set_untrusted_text


def test_untrusted_prefixes_are_escaped_as_text(tmp_path) -> None:
    """P0-EXP-007J writes formula-like external values as ordinary strings."""
    path = tmp_path / "safe.xlsx"
    values = ("=1+1", "+cmd", "-2+3", "@SUM(A1:A2)")
    book = Workbook()
    for row, value in enumerate(values, start=1):
        set_untrusted_text(book.active.cell(row=row, column=1), value)
    book.save(path)
    book.close()

    reopened = load_workbook(path, data_only=False)
    try:
        for row, expected in enumerate(values, start=1):
            cell = reopened.active.cell(row=row, column=1)
            assert cell.data_type == "s"
            assert cell.value == expected
    finally:
        reopened.close()
